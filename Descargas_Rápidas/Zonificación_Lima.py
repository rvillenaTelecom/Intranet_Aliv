"""
ALIV TELECOM — Clasificador de Zonas KML para WinForce
=======================================================
Cruza cada registro de WinForce con los polígonos del KML
usando las coordenadas exactas del cliente.

Resultado: agrega columna Zona_KML a WinForce con:
  - Zona P2 (401)     → score mínimo 401
  - No Venta          → bloqueada, será desaprobada
  - Sin modificación  → score mínimo 201
  - Sin coordenada    → no tiene lat/lon para cruzar

Cómo usar:
  1. Pon este script en la misma carpeta que los archivos
  2. Ajusta las rutas en CONFIGURACIÓN
  3. Ejecuta: python clasificar_winforce.py
"""

import sys
import xml.etree.ElementTree as ET
from shapely.geometry import Polygon, Point
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

import os
from db_config import upload_to_sql, upload_incremental_to_sql

# ─────────────────────────────────────────
#  CONFIGURACIÓN Y BÚSQUEDA
# ─────────────────────────────────────────

def buscar_winforce():
    # Rutas probables
    posibles = [
        "Winforce_Lima.xlsx",
        os.path.join("descargas_winforce_Dept", "Winforce_Lima.xlsx"),
        os.path.join("..", "descargas_winforce_Dept", "Winforce_Lima.xlsx")
    ]
    for p in posibles:
        if os.path.exists(p):
            return p
    return "Winforce_Lima.xlsx" # Fallback

RUTA_WINFORCE  = buscar_winforce()
RUTA_KML       = r"Parametros_ventas.kml"
# RUTA_SALIDA se ha eliminado para modificar el archivo original directamente

# ─────────────────────────────────────────

def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

def cargar_poligonos(ruta_kml):
    """Carga los polígonos del KML separados por tipo."""
    log("Leyendo KML...")
    tree = ET.parse(ruta_kml)
    root = tree.getroot()
    ns = '{http://www.opengis.net/kml/2.2}'

    polys_nv = []   # No Venta (rojo)
    polys_p2 = []   # Preferente 2 (celeste)

    for folder in root.iter(f'{ns}Folder'):
        fn = folder.find(f'{ns}name')
        capa = fn.text.strip() if (fn is not None and fn.text) else ''
        es_nv = 'BLOQUEADO' in capa.upper()

        for pm in folder.iter(f'{ns}Placemark'):
            coords_el = pm.find(f'.//{ns}coordinates')
            if coords_el is None or not coords_el.text:
                continue
            points = []
            for pt in coords_el.text.strip().split():
                parts = pt.split(',')
                if len(parts) >= 2:
                    try:
                        points.append((float(parts[0]), float(parts[1])))
                    except:
                        pass
            if len(points) < 3:
                continue
            try:
                poly = Polygon(points)
                if not poly.is_valid:
                    poly = poly.buffer(0)
                if poly.is_valid:
                    if es_nv:
                        polys_nv.append(poly)
                    else:
                        polys_p2.append(poly)
            except:
                pass

    log(f"  Poligonos P2 (celeste): {len(polys_p2)}")
    log(f"  Poligonos No Venta (rojo): {len(polys_nv)}")
    return polys_p2, polys_nv


def clasificar_punto(lat, lon, polys_p2, polys_nv):
    """
    Determina la zona de un punto geográfico.
    Primero verifica No Venta (más restrictivo),
    luego P2, luego Sin modificación por descarte.
    """
    try:
        pt = Point(lon, lat)
        for poly in polys_nv:
            if poly.contains(pt):
                return 'No Venta'
        for poly in polys_p2:
            if poly.contains(pt):
                return 'Zona P2 (401)'
        return 'Sin modificación (201)'
    except:
        return 'Sin coordenada'


def main():
    print("=" * 60)
    print("  ALIV TELECOM - Clasificador de Zonas WinForce")
    print("  Fuente: PARAMETROS_VENTAS.kml")
    print("=" * 60)

    # Cargar polígonos
    polys_p2, polys_nv = cargar_poligonos(RUTA_KML)

    # Cargar WinForce
    log(f"Leyendo WinForce desde: {RUTA_WINFORCE}")
    if not os.path.exists(RUTA_WINFORCE):
        log(f"ERROR: No se encontro el archivo {RUTA_WINFORCE}")
        return
    
    # Validar que el archivo no es el output procesado anterior (hoja Resumen como primera hoja)
    xl = pd.ExcelFile(RUTA_WINFORCE)
    if 'Resumen' in xl.sheet_names and xl.sheet_names[0] == 'Resumen':
        log("ERROR: El archivo es el output procesado anterior, no una descarga fresca de WinForce.")
        log("  Ejecuta primero WinforceLima2026.py para obtener datos actualizados.")
        sys.exit(1)

    df = pd.read_excel(RUTA_WINFORCE)
    log(f"  Registros cargados: {len(df):,}")

    # Buscar columnas de coordenadas sin importar si vienen en mayuscula o minuscula
    cols_lower = {c.lower(): c for c in df.columns}
    col_lat = cols_lower.get('latitud')
    col_lon = cols_lower.get('longitud')

    if not col_lat or not col_lon:
        log("AVISO: El archivo no tiene columnas latitud/longitud. Todos los registros se marcan como 'Sin coordenada'.")
        df['Zona_KML'] = 'Sin coordenada'
        df['Score_Minimo_KML'] = None
    else:
        # Limpiar coordenadas
        df[col_lat] = pd.to_numeric(df[col_lat], errors='coerce')
        df[col_lon] = pd.to_numeric(df[col_lon], errors='coerce')

        con_coords = df[col_lat].notna() & df[col_lon].notna()
        log(f"  Con coordenadas: {con_coords.sum():,}")
        log(f"  Sin coordenadas: {(~con_coords).sum():,}")

        # Clasificar — solo Lima (coordenadas dentro del bbox del KML)
        lima_mask = (
            df[col_lat].between(-13, -11) &
            df[col_lon].between(-77.5, -76.5)
        )
        log(f"  En bbox de Lima: {lima_mask.sum():,}")
        log("Clasificando registros por zona (puede tardar 1-2 min)...")

        df['Zona_KML'] = 'Sin coordenada'
        df.loc[lima_mask, 'Zona_KML'] = df[lima_mask].apply(
            lambda r: clasificar_punto(r[col_lat], r[col_lon], polys_p2, polys_nv),
            axis=1
        )
        # Los que tienen coordenadas pero fuera de Lima = Sin modificación
        fuera_lima = con_coords & ~lima_mask
        df.loc[fuera_lima, 'Zona_KML'] = 'Sin modificación (201)'

        df['Score_Minimo_KML'] = df['Zona_KML'].map({
            'Zona P2 (401)':          401,
            'No Venta':               0,
            'Sin modificación (201)': 201,
            'Sin coordenada':         None,
        })

    # Resumen
    print()
    print("=" * 60)
    print("  RESULTADO DE CLASIFICACION")
    print("=" * 60)
    total = len(df)
    for zona, n in df['Zona_KML'].value_counts().items():
        pct = n / total * 100
        print(f"  {zona:<28} {n:>6,}  ({pct:.1f}%)")

    print()
    # No Venta por distrito
    nv = df[df['Zona_KML'] == 'No Venta']
    if len(nv) > 0:
        print("  Top distritos en Zona No Venta:")
        for dist, n in nv['Distrito'].value_counts().head(8).items():
            print(f"    {dist:<30} {n:>5,}")

    print()
    # P2 sin score (potencial riesgo)
    p2 = df[df['Zona_KML'] == 'Zona P2 (401)']
    print(f"  Registros en Zona P2: {len(p2):,}")
    print("=" * 60)

    # Exportar Excel (Sobrescribir original)
    log(f"Modificando archivo original: {RUTA_WINFORCE}...")

    ORANGE = 'F47920'; DARK = '1A1A2E'; WHITE = 'FFFFFF'; GRIS = 'F5F5F3'
    s = Side(style='thin', color='CCCCCC')
    brd = Border(left=s, right=s, top=s, bottom=s)

    color_zona = {
        'Zona P2 (401)':          'DAEEF8',
        'No Venta':               'FCEBEB',
        'Sin modificación (201)': 'F5F5F3',
        'Sin coordenada':         'EDEDEB',
    }

    wb = Workbook()

    # ── Hoja 1: Resumen ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Resumen'
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells('A1:E1')
    ws1['A1'] = f'Clasificación de Zonas KML — WinForce Lima · {datetime.now().strftime("%d/%m/%Y")}'
    ws1['A1'].font = Font(bold=True, size=12, color=WHITE, name='Arial')
    ws1['A1'].fill = PatternFill('solid', start_color=DARK)
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 26

    hdrs = ['Zona KML', 'Registros', '% del total', 'Score Mínimo', 'Acción']
    for i, h in enumerate(hdrs):
        cell = ws1.cell(row=2, column=i+1, value=h)
        cell.font = Font(bold=True, color=WHITE, size=10, name='Arial')
        cell.fill = PatternFill('solid', start_color=ORANGE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = brd
    ws1.row_dimensions[2].height = 18

    acciones = {
        'Zona P2 (401)':          'Verificar score ≥ 401 antes de validar',
        'No Venta':               '🚫 CRÍTICO — será desaprobada en validación',
        'Sin modificación (201)': 'Sin restricción adicional — score 201',
        'Sin coordenada':         'Revisar coordenadas del registro',
    }

    for zona, n in df['Zona_KML'].value_counts().items():
        r = ws1.max_row + 1
        pct = n / total
        bg = color_zona.get(zona, GRIS)
        vals = [zona, n, pct, df[df['Zona_KML']==zona]['Score_Minimo_KML'].iloc[0], acciones.get(zona,'')]
        for j, v in enumerate(vals):
            cell = ws1.cell(row=r, column=j+1, value=v)
            cell.font = Font(size=10, name='Arial', color=DARK)
            cell.fill = PatternFill('solid', start_color=bg)
            cell.alignment = Alignment(vertical='center', horizontal='left' if j in [0,4] else 'center')
            cell.border = brd
        ws1.cell(row=r, column=3).number_format = '0.0%'
        ws1.row_dimensions[r].height = 16

    for i, w in enumerate([28, 12, 12, 14, 40]):
        ws1.column_dimensions[get_column_letter(i+1)].width = w

    # ── Hoja 2: No Venta ─────────────────────────────────────────
    ws2 = wb.create_sheet('No Venta')
    ws2.sheet_view.showGridLines = False

    df_nv = df[df['Zona_KML'] == 'No Venta'].copy()
    ws2.merge_cells(f'A1:{get_column_letter(len(df.columns))}1')
    ws2['A1'] = f'Registros en Zona No Venta — {len(df_nv):,} registros — CRÍTICO'
    ws2['A1'].font = Font(bold=True, size=12, color=WHITE, name='Arial')
    ws2['A1'].fill = PatternFill('solid', start_color='E24B4A')
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 26

    for i, h in enumerate(df.columns):
        cell = ws2.cell(row=2, column=i+1, value=h)
        cell.font = Font(bold=True, color=WHITE, size=9, name='Arial')
        cell.fill = PatternFill('solid', start_color=ORANGE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = brd
    ws2.row_dimensions[2].height = 16

    for _, row in df_nv.iterrows():
        r = ws2.max_row + 1
        for j, v in enumerate(row):
            cell = ws2.cell(row=r, column=j+1, value=v)
            cell.font = Font(size=9, name='Arial', color=DARK)
            cell.fill = PatternFill('solid', start_color='FCEBEB')
            cell.border = brd
        ws2.row_dimensions[r].height = 13

    ws2.freeze_panes = 'A3'
    ws2.auto_filter.ref = f'A2:{get_column_letter(len(df.columns))}2'

    # ── Hoja 3: Data completa ────────────────────────────────────
    ws3 = wb.create_sheet('Data Completa')
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells(f'A1:{get_column_letter(len(df.columns))}1')
    ws3['A1'] = f'WinForce Lima con Zona KML — {len(df):,} registros — {datetime.now().strftime("%d/%m/%Y")}'
    ws3['A1'].font = Font(bold=True, size=12, color=WHITE, name='Arial')
    ws3['A1'].fill = PatternFill('solid', start_color=DARK)
    ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 26

    for i, h in enumerate(df.columns):
        cell = ws3.cell(row=2, column=i+1, value=h)
        cell.font = Font(bold=True, color=WHITE, size=9, name='Arial')
        cell.fill = PatternFill('solid', start_color=ORANGE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = brd
    ws3.row_dimensions[2].height = 16

    for _, row in df.iterrows():
        r = ws3.max_row + 1
        bg = color_zona.get(row.get('Zona_KML', ''), GRIS)
        for j, v in enumerate(row):
            cell = ws3.cell(row=r, column=j+1, value=v)
            cell.font = Font(size=9, name='Arial', color=DARK)
            cell.fill = PatternFill('solid', start_color=bg)
            cell.border = brd
        ws3.row_dimensions[r].height = 13

    ws3.freeze_panes = 'A3'
    ws3.auto_filter.ref = f'A2:{get_column_letter(len(df.columns))}2'

    wb.save(RUTA_WINFORCE)
    log(f"Cambios aplicados correctamente en: {RUTA_WINFORCE}")

    print()
    print("=" * 60)
    print("  [OK] PROCESO COMPLETADO - ARCHIVO ACTUALIZADO")
    print(f"  Ruta: {RUTA_WINFORCE}")
    print("  Hojas actualizadas:")
    print("    1. Resumen — conteo por zona")
    print("    2. No Venta — registros críticos")
    print("    3. Data Completa — todos los registros con Zona_KML")
    print()
    print("  Nota: El archivo original ha sido modificado con la")
    print("  nueva clasificación. Ya puedes usarlo en Power BI.")
    print("=" * 60)
    
    # NUEVO: Cargar a SQL Server después de zonificar
    try:
        print("\n  Iniciando carga a SQL Server (winforce_lima)...")
        # El DataFrame 'df' ya contiene las nuevas columnas de zonificación
        
        incremental = "--incremental" in sys.argv
        if incremental:
            # En modo incremental, borramos los últimos 7 días y subimos lo nuevo
            upload_incremental_to_sql(df, "winforce_lima", "Fecha de registro")
        else:
            # En modo completo, reemplazamos toda la tabla
            upload_to_sql(df, "winforce_lima")
            
    except Exception as sql_e:
        print(f"  [SQL] Error al cargar: {sql_e}")
        
    print("\nZonificacion exitosa")


if __name__ == "__main__":
    main()