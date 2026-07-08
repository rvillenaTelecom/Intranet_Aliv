"""
Forecasting/forecasting_ventas.py
===================================
Predicción (forecasting) de altas y ventas diarias y mensuales para Lima.
Usa Holt-Winters Exponential Smoothing sobre datos históricos 2024-2026.

Métricas predichas:
  • Altas instaladas (Estado orden = 'Ejecutada')
  • Ventas registradas (con Plan asignado)
  — En total, Vertical y Horizontal —

Uso:
    python forecasting_ventas.py              # 60 días de forecast
    python forecasting_ventas.py --dias 90    # 90 días de forecast

Salida:
    Forecast_Aliv_YYYY-MM-DD.xlsx
"""

import sys
import os
import argparse
import warnings
import logging
from datetime import date, datetime, timedelta

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import io

from statsmodels.tsa.holtwinters import ExponentialSmoothing

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')
logging.getLogger('statsmodels').setLevel(logging.ERROR)

# ── Ruta al módulo de base de datos ─────────────────────────────
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'Descargas_Rápidas', 'Pythons'))
from db_config import get_engine
import sqlalchemy as sa

# ── Constantes ───────────────────────────────────────────────────
VERTICAL_TIPOS = {'Condominio/Edificio', 'C/E Habilitado'}

C_NARANJA = '#F47920'
C_AZUL    = '#2E86AB'
C_VERDE   = '#1D9E75'
C_VIOLETA = '#7B2D8B'
C_GRIS    = '#AAAAAA'
C_ROJO    = '#E24B4A'


# ═══════════════════════════════════════════════════════════════
# 1. EXTRACCIÓN DE DATOS
# ═══════════════════════════════════════════════════════════════

def cargar_historico():
    """
    Carga altas y ventas diarias de las 3 tablas (2024, 2025, 2026).
    Devuelve un dict con DataFrames diarios por segmento.
    """
    engine = get_engine()
    hoy    = pd.Timestamp(date.today())

    altas_raw  = []
    ventas_raw = []

    tablas = [
        ('winforce_lima_2024', 2024),
        ('winforce_lima_2025', 2025),
        ('winforce_lima',      2026),
    ]

    with engine.connect() as conn:
        insp = sa.inspect(engine)
        for tabla, anio in tablas:
            if not insp.has_table(tabla, schema='dbo'):
                print(f"  [OMITIDO] Tabla {tabla} no existe.")
                continue

            # ── Altas: Fecha programación + Tipo de domicilio ──
            try:
                df_a = pd.read_sql(sa.text(f"""
                    SELECT
                        TRY_CONVERT(DATE, LEFT([Fecha programación], 10), 105) AS fecha,
                        [Tipo de domicilio] AS tipo_dom
                    FROM [{tabla}]
                    WHERE [Estado orden] = 'Ejecutada'
                      AND [Fecha programación] IS NOT NULL
                """), conn)
                df_a['fecha'] = pd.to_datetime(df_a['fecha'])
                df_a['vertical'] = df_a['tipo_dom'].isin(VERTICAL_TIPOS)
                altas_raw.append(df_a)
                print(f"  {tabla}: {len(df_a):,} altas")
            except Exception as e:
                print(f"  [WARN] Altas de {tabla}: {e}")

            # ── Ventas: Fecha de registro + Tipo de domicilio ──
            try:
                df_v = pd.read_sql(sa.text(f"""
                    SELECT
                        CAST([Fecha de registro] AS DATE) AS fecha,
                        [Tipo de domicilio] AS tipo_dom
                    FROM [{tabla}]
                    WHERE [Plan] IS NOT NULL AND [Plan] <> ''
                      AND [Fecha de registro] IS NOT NULL
                """), conn)
                df_v['fecha'] = pd.to_datetime(df_v['fecha'])
                df_v['vertical'] = df_v['tipo_dom'].isin(VERTICAL_TIPOS)
                ventas_raw.append(df_v)
                print(f"  {tabla}: {len(df_v):,} ventas")
            except Exception as e:
                print(f"  [WARN] Ventas de {tabla}: {e}")

    def _serie_diaria(dfs, segmento='total'):
        """Agrega filas a una serie diaria completa (sin huecos)."""
        df = pd.concat(dfs, ignore_index=True)
        df = df[(df['fecha'] >= '2024-01-01') & (df['fecha'] <= hoy)]
        df = df.dropna(subset=['fecha'])

        if segmento == 'vertical':
            df = df[df['vertical']]
        elif segmento == 'horizontal':
            df = df[~df['vertical']]

        daily = (df.groupby('fecha').size()
                   .reset_index(name='y')
                   .rename(columns={'fecha': 'ds'}))

        # Rellenamos días sin actividad con 0 para evitar huecos
        idx = pd.date_range(daily['ds'].min(), hoy, freq='D')
        daily = (daily.set_index('ds')
                      .reindex(idx, fill_value=0)
                      .reset_index()
                      .rename(columns={'index': 'ds'}))
        return daily

    result = {}
    if altas_raw:
        result['altas_total']    = _serie_diaria(altas_raw, 'total')
        result['altas_vert']     = _serie_diaria(altas_raw, 'vertical')
        result['altas_horiz']    = _serie_diaria(altas_raw, 'horizontal')

    if ventas_raw:
        result['ventas_total']   = _serie_diaria(ventas_raw, 'total')
        result['ventas_vert']    = _serie_diaria(ventas_raw, 'vertical')
        result['ventas_horiz']   = _serie_diaria(ventas_raw, 'horizontal')

    return result


# ═══════════════════════════════════════════════════════════════
# 2. MODELO DE FORECASTING — HOLT-WINTERS
# ═══════════════════════════════════════════════════════════════

def entrenar_hw(df_diario, dias_forecast=60):
    """
    Ajusta Holt-Winters con estacionalidad semanal (periodo=7).
    Devuelve (serie_historica, serie_forecast, conf_intervals).
    """
    if df_diario is None or len(df_diario) < 21:
        return None

    df = df_diario.copy()
    df['ds'] = pd.to_datetime(df['ds'])
    df = df.sort_values('ds').set_index('ds')

    serie = df['y'].astype(float)
    serie = serie.clip(lower=0)

    # Holt-Winters: tendencia aditiva + estacionalidad aditiva semanal
    modelo = ExponentialSmoothing(
        serie,
        trend='add',
        seasonal='add',
        seasonal_periods=7,
        initialization_method='estimated',
    )
    fit = modelo.fit(optimized=True, use_brute=False)

    # Forecast
    pred = fit.forecast(dias_forecast)
    pred = pred.clip(lower=0)

    # Intervalo de confianza aproximado (±1.96 * RMSE del modelo)
    residuos = fit.resid
    rmse = np.sqrt(np.mean(residuos ** 2))
    lower = (pred - 1.96 * rmse).clip(lower=0)
    upper = pred + 1.96 * rmse

    df_forecast = pd.DataFrame({
        'ds':    pred.index,
        'yhat':  pred.values.round(1),
        'lower': lower.values.round(1),
        'upper': upper.values.round(1),
    })

    return df_forecast, fit


def forecast_mensual(df_hist, df_forecast):
    """Agrega histórico + forecast a nivel mensual."""
    hoy = pd.Timestamp(date.today())

    hist = df_hist.copy()
    hist['ds'] = pd.to_datetime(hist['ds'])
    hist['mes'] = hist['ds'].dt.to_period('M')
    hist_m = hist.groupby('mes')['y'].sum().reset_index()
    hist_m.columns = ['mes', 'real']

    fc = df_forecast.copy()
    fc['ds'] = pd.to_datetime(fc['ds'])
    fc['mes'] = fc['ds'].dt.to_period('M')
    fc_m = fc.groupby('mes')['yhat'].sum().reset_index()
    fc_m.columns = ['mes', 'forecast']

    merged = hist_m.merge(fc_m, on='mes', how='outer').sort_values('mes')

    # Mes actual: suma real + forecast de días restantes
    mes_actual = pd.Period(hoy, 'M')
    mask_actual = merged['mes'] == mes_actual
    if mask_actual.any():
        real_act = merged.loc[mask_actual, 'real'].fillna(0).values[0]
        fc_act   = merged.loc[mask_actual, 'forecast'].fillna(0).values[0]
        merged.loc[mask_actual, 'proyeccion'] = real_act + fc_act
    else:
        merged['proyeccion'] = merged['real']

    merged.loc[merged['mes'] < mes_actual, 'proyeccion'] = merged.loc[merged['mes'] < mes_actual, 'real']
    merged.loc[merged['mes'] > mes_actual, 'proyeccion'] = merged.loc[merged['mes'] > mes_actual, 'forecast']

    merged['tipo'] = merged['mes'].apply(
        lambda m: 'Histórico' if m < mes_actual else ('Actual' if m == mes_actual else 'Forecast')
    )
    return merged


# ═══════════════════════════════════════════════════════════════
# 3. GRÁFICOS
# ═══════════════════════════════════════════════════════════════

def _buf(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=130, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    return buf


def grafico_diario(df_hist, df_forecast, titulo, color_hist, color_pred):
    hoy = pd.Timestamp(date.today())

    fig, ax = plt.subplots(figsize=(13, 4.2))
    fig.patch.set_facecolor('#FAFAFA')
    ax.set_facecolor('#FAFAFA')

    # Histórico — últimos 12 meses para no saturar
    df_h = df_hist[df_hist['ds'] >= hoy - pd.Timedelta(days=365)].copy()
    ax.plot(df_h['ds'], df_h['y'], color=color_hist, linewidth=1.1,
            alpha=0.75, label='Real', zorder=3)

    # Media móvil 7 días sobre el histórico
    mm7 = df_h.set_index('ds')['y'].rolling(7, center=True).mean()
    ax.plot(mm7.index, mm7.values, color=color_hist, linewidth=2,
            alpha=0.9, label='Media móvil 7d', zorder=4)

    # Forecast
    ax.plot(df_forecast['ds'], df_forecast['yhat'],
            color=color_pred, linewidth=2.2, linestyle='--',
            label='Predicción', zorder=5)
    ax.fill_between(df_forecast['ds'], df_forecast['lower'], df_forecast['upper'],
                    color=color_pred, alpha=0.15, label='Intervalo 95%')

    ax.axvline(hoy, color='#777777', linestyle=':', linewidth=1.2, alpha=0.8)
    ax.text(hoy, ax.get_ylim()[1] * 0.97, '  Hoy', color='#777777', fontsize=8)

    ax.set_title(titulo, fontsize=12, fontweight='bold', pad=10, color='#1C1C1E')
    ax.set_ylabel('Cantidad / día', fontsize=9)
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%b %y'))
    ax.xaxis.set_major_locator(mdates.MonthLocator())
    plt.xticks(rotation=30, fontsize=8)
    ax.yaxis.set_tick_params(labelsize=8)
    ax.legend(fontsize=8, loc='upper left', framealpha=0.7)
    ax.grid(axis='y', alpha=0.3, linestyle='--')
    ax.spines[['top', 'right']].set_visible(False)
    plt.tight_layout()
    return _buf(fig)


def grafico_mensual(df_mensual, titulo, color):
    fig, ax = plt.subplots(figsize=(13, 4.5))
    fig.patch.set_facecolor('#FAFAFA')
    ax.set_facecolor('#FAFAFA')

    meses_str = [str(m) for m in df_mensual['mes']]
    tipos      = df_mensual['tipo'].tolist()
    proyeccion = df_mensual['proyeccion'].fillna(0).tolist()
    real       = df_mensual['real'].fillna(0).tolist()

    colores = []
    hatch   = []
    for t in tipos:
        if t == 'Histórico':
            colores.append(color)
            hatch.append('')
        elif t == 'Actual':
            colores.append(color)
            hatch.append('//')
        else:
            colores.append(color)
            hatch.append('xx')

    bars = ax.bar(meses_str, proyeccion, color=colores, alpha=0.82,
                  edgecolor=color)
    for bar, h in zip(bars, hatch):
        bar.set_hatch(h)

    for i, (val, tipo) in enumerate(zip(proyeccion, tipos)):
        ax.text(i, val + max(proyeccion) * 0.01,
                f'{int(round(val)):,}',
                ha='center', va='bottom', fontsize=7.5, fontweight='bold',
                color='#333333')

    from matplotlib.patches import Patch
    legend_elements = [
        Patch(facecolor=color, alpha=0.82, label='Histórico real'),
        Patch(facecolor=color, alpha=0.82, hatch='//', label='Mes actual (parcial+forecast)'),
        Patch(facecolor=color, alpha=0.82, hatch='xx', label='Forecast'),
    ]
    ax.legend(handles=legend_elements, fontsize=8, loc='upper left', framealpha=0.7)
    ax.set_title(titulo, fontsize=12, fontweight='bold', pad=10, color='#1C1C1E')
    ax.set_ylabel('Cantidad mensual', fontsize=9)
    plt.xticks(rotation=45, fontsize=8)
    ax.yaxis.set_tick_params(labelsize=8)
    ax.grid(axis='y', alpha=0.3, linestyle='--')
    ax.spines[['top', 'right']].set_visible(False)
    plt.tight_layout()
    return _buf(fig)


def grafico_comparativo(resultados, tipo_metrica):
    """Gráfico de barras apiladas por mes: Vertical vs Horizontal."""
    key_vert  = f'{tipo_metrica}_vert'
    key_horiz = f'{tipo_metrica}_horiz'

    if key_vert not in resultados or key_horiz not in resultados:
        return None

    def mensual_proy(key):
        hist, fc = resultados[key]
        m = forecast_mensual(hist, fc)
        return m.set_index('mes')['proyeccion'].fillna(0)

    sv = mensual_proy(key_vert)
    sh = mensual_proy(key_horiz)
    meses = sorted(set(sv.index) | set(sh.index))
    v_vals = [sv.get(m, 0) for m in meses]
    h_vals = [sh.get(m, 0) for m in meses]
    meses_str = [str(m) for m in meses]

    fig, ax = plt.subplots(figsize=(13, 4.5))
    fig.patch.set_facecolor('#FAFAFA')
    ax.set_facecolor('#FAFAFA')

    ax.bar(meses_str, v_vals, label='Vertical', color=C_NARANJA, alpha=0.85)
    ax.bar(meses_str, h_vals, bottom=v_vals, label='Horizontal', color=C_AZUL, alpha=0.85)

    for i, (vv, hv) in enumerate(zip(v_vals, h_vals)):
        total = vv + hv
        if total > 0:
            ax.text(i, total + max(v for v in [vv+hv for vv, hv in zip(v_vals, h_vals)]) * 0.01,
                    f'{int(round(total)):,}', ha='center', va='bottom', fontsize=7.5, fontweight='bold')

    titulo = 'Altas mensuales' if tipo_metrica == 'altas' else 'Ventas mensuales'
    ax.set_title(f'{titulo} — Vertical vs Horizontal (histórico + forecast)',
                 fontsize=12, fontweight='bold', pad=10, color='#1C1C1E')
    ax.set_ylabel('Cantidad mensual', fontsize=9)
    plt.xticks(rotation=45, fontsize=8)
    ax.yaxis.set_tick_params(labelsize=8)
    ax.legend(fontsize=9, loc='upper left', framealpha=0.7)
    ax.grid(axis='y', alpha=0.3, linestyle='--')
    ax.spines[['top', 'right']].set_visible(False)
    plt.tight_layout()
    return _buf(fig)


# ═══════════════════════════════════════════════════════════════
# 4. EXCEL DE SALIDA
# ═══════════════════════════════════════════════════════════════

def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color.replace('#', ''))

def _font(bold=False, color='000000', size=9):
    return Font(bold=bold, color=color, size=size)

def _border():
    thin = Side(style='thin', color='DDDDDD')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _encabezado(ws, fila_idx, color_hex):
    for cell in ws[fila_idx]:
        if cell.value is not None:
            cell.fill    = _fill(color_hex)
            cell.font    = _font(bold=True, color='FFFFFF', size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border  = _border()

def _datos_row(ws, fila_idx, es_forecast=False):
    color_bg = 'FFF3E0' if es_forecast else 'FFFFFF'
    for cell in ws[fila_idx]:
        if cell.value is not None:
            cell.fill   = _fill(color_bg)
            cell.font   = _font(size=9)
            cell.border = _border()
            cell.alignment = Alignment(horizontal='center')


def sheet_resumen(wb, resultados, dias_forecast):
    """Hoja principal: tabla mensual consolidada de todos los segmentos."""
    ws = wb.create_sheet("Resumen Mensual", 0)
    hoy = pd.Timestamp(date.today())

    # Título
    ws.merge_cells('A1:J1')
    ws['A1'] = f"FORECAST VENTAS ALIV — Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A1'].font = _font(bold=True, color='7B2D8B', size=13)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:J2')
    ws['A2'] = f"Modelo: Holt-Winters Exponential Smoothing · Horizonte: {dias_forecast} días · Datos: Ene 2024 – {hoy.strftime('%b %Y')}"
    ws['A2'].font = _font(color='888888', size=9)
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.row_dimensions[3].height = 6  # espacio

    # Encabezados
    headers = [
        'Mes', 'Tipo',
        'Altas Total', 'Altas Vertical', 'Altas Horizontal',
        'Ventas Total', 'Ventas Vertical', 'Ventas Horizontal',
    ]
    ws.append([''] * len(headers))  # fila 4 vacía
    ws.append(headers)
    _encabezado(ws, 5, '1C1C1E')

    # Recopilar datos mensuales por segmento
    series_m = {}
    for key in ['altas_total', 'altas_vert', 'altas_horiz', 'ventas_total', 'ventas_vert', 'ventas_horiz']:
        if key in resultados and resultados[key] is not None:
            hist, fc = resultados[key]
            m = forecast_mensual(hist, fc)
            series_m[key] = m.set_index('mes')

    # Todos los meses disponibles
    meses = sorted(set().union(*[set(df.index) for df in series_m.values()]))

    for mes in meses:
        tipo = 'Histórico'
        mp = pd.Period(hoy, 'M')
        if mes == mp:
            tipo = 'Mes actual'
        elif mes > mp:
            tipo = 'Forecast'

        def val(key):
            if key not in series_m or mes not in series_m[key].index:
                return ''
            v = series_m[key].loc[mes, 'proyeccion']
            return int(round(v)) if pd.notna(v) else ''

        ws.append([
            str(mes), tipo,
            val('altas_total'), val('altas_vert'), val('altas_horiz'),
            val('ventas_total'), val('ventas_vert'), val('ventas_horiz'),
        ])
        _datos_row(ws, ws.max_row, es_forecast=(tipo == 'Forecast'))

    # Anchos de columna
    for col, w in zip(['A','B','C','D','E','F','G','H'], [10, 12, 12, 14, 16, 12, 14, 16]):
        ws.column_dimensions[col].width = w

    # Congelar encabezado
    ws.freeze_panes = 'A6'
    return ws


def sheet_diario_excel(wb, nombre_hoja, df_hist, df_forecast, titulo_grafico,
                        color_hist, color_pred, color_enc):
    ws = wb.create_sheet(nombre_hoja)
    hoy = pd.Timestamp(date.today())

    hist_lookup = df_hist.set_index('ds')['y'].to_dict()

    headers = ['Fecha', 'Real', 'Predicción', 'Límite Inf.', 'Límite Sup.', 'Tipo']
    ws.append(headers)
    _encabezado(ws, 1, color_enc)

    for _, row in df_hist.iterrows():
        ws.append([
            row['ds'].strftime('%Y-%m-%d'),
            int(row['y']),
            '', '', '',
            'Histórico',
        ])
        _datos_row(ws, ws.max_row)

    for _, row in df_forecast.iterrows():
        ws.append([
            row['ds'].strftime('%Y-%m-%d'),
            '',
            round(row['yhat'], 1),
            round(row['lower'], 1),
            round(row['upper'], 1),
            'Forecast',
        ])
        _datos_row(ws, ws.max_row, es_forecast=True)

    for col, w in zip(['A','B','C','D','E','F'], [12, 8, 10, 12, 12, 10]):
        ws.column_dimensions[col].width = w

    ws.freeze_panes = 'A2'

    # Gráfico
    buf = grafico_diario(df_hist, df_forecast, titulo_grafico, color_hist, color_pred)
    img = XLImage(buf)
    img.width  = 780
    img.height = 270
    ws.add_image(img, 'H2')
    return ws


def sheet_mensual_excel(wb, nombre_hoja, df_hist, df_forecast, titulo_grafico, color):
    ws = wb.create_sheet(nombre_hoja)
    df_m = forecast_mensual(df_hist, df_forecast)

    headers = ['Mes', 'Real', 'Forecast', 'Proyección', 'Tipo']
    ws.append(headers)
    _encabezado(ws, 1, color.replace('#', ''))

    for _, row in df_m.iterrows():
        ws.append([
            str(row['mes']),
            int(round(row['real'])) if pd.notna(row['real']) else '',
            int(round(row['forecast'])) if pd.notna(row.get('forecast')) else '',
            int(round(row['proyeccion'])) if pd.notna(row['proyeccion']) else '',
            row['tipo'],
        ])
        _datos_row(ws, ws.max_row, es_forecast=(row['tipo'] == 'Forecast'))

    for col, w in zip(['A','B','C','D','E'], [10, 10, 10, 11, 10]):
        ws.column_dimensions[col].width = w

    ws.freeze_panes = 'A2'

    buf = grafico_mensual(df_m, titulo_grafico, color)
    img = XLImage(buf)
    img.width  = 780
    img.height = 290
    ws.add_image(img, 'G2')
    return ws


def sheet_comparativo_excel(wb, resultados, tipo_metrica):
    nombre = 'Altas por Segmento' if tipo_metrica == 'altas' else 'Ventas por Segmento'
    ws = wb.create_sheet(nombre)
    buf = grafico_comparativo(resultados, tipo_metrica)
    if buf:
        img = XLImage(buf)
        img.width  = 900
        img.height = 300
        ws.add_image(img, 'A2')
    return ws


# ═══════════════════════════════════════════════════════════════
# 5. ORQUESTADOR PRINCIPAL
# ═══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description='Forecasting de ventas Aliv')
    parser.add_argument('--dias', type=int, default=60, help='Días de predicción (default: 60)')
    args = parser.parse_args()
    dias_forecast = args.dias

    print(f"\n{'='*55}")
    print(f"  FORECASTING VENTAS ALIV — {dias_forecast} días de predicción")
    print(f"{'='*55}\n")

    # 1. Cargar datos
    print("[ 1/4 ] Cargando datos históricos...")
    datos = cargar_historico()
    print(f"  Total registros cargados: {sum(len(v) for v in datos.values()):,} filas diarias\n")

    # 2. Entrenar modelos
    print("[ 2/4 ] Entrenando modelos Holt-Winters...")
    resultados = {}
    configs = [
        ('altas_total',  'Altas Lima Total',       C_VERDE,   C_VIOLETA),
        ('altas_vert',   'Altas Vertical',          C_NARANJA, C_ROJO),
        ('altas_horiz',  'Altas Horizontal',        C_AZUL,    C_VIOLETA),
        ('ventas_total', 'Ventas Lima Total',        C_VERDE,   C_VIOLETA),
        ('ventas_vert',  'Ventas Vertical',          C_NARANJA, C_ROJO),
        ('ventas_horiz', 'Ventas Horizontal',        C_AZUL,    C_VIOLETA),
    ]

    for key, nombre, c_hist, c_pred in configs:
        if key not in datos:
            print(f"  {nombre}: sin datos.")
            continue
        try:
            res = entrenar_hw(datos[key], dias_forecast)
            if res:
                df_fc, fit = res
                resultados[key] = (datos[key], df_fc)
                # Resumen de forecast próximo mes
                prox_mes = (pd.Timestamp(date.today()) + pd.Timedelta(days=30))
                fc_30 = df_fc[df_fc['ds'] <= prox_mes]['yhat'].sum()
                print(f"  {nombre:<25} -> proximos 30 dias: {fc_30:,.0f}  |  proximos {dias_forecast} dias: {df_fc['yhat'].sum():,.0f}")
        except Exception as e:
            print(f"  [ERROR] {nombre}: {e}")

    print()

    # 3. Generar Excel
    print("[ 3/4 ] Generando Excel con tablas y gráficos...")
    wb = Workbook()
    wb.remove(wb.active)

    sheet_resumen(wb, resultados, dias_forecast)

    if 'altas_total' in resultados:
        h, fc = resultados['altas_total']
        sheet_diario_excel(wb, 'Altas Diarias', h, fc,
                           'Altas Instaladas — Lima Total (diario + forecast)',
                           C_VERDE, C_VIOLETA, '1D9E75')
        sheet_mensual_excel(wb, 'Altas Mensuales', h, fc,
                            'Altas Mensuales — Lima Total',
                            C_VERDE)

    if 'ventas_total' in resultados:
        h, fc = resultados['ventas_total']
        sheet_diario_excel(wb, 'Ventas Diarias', h, fc,
                           'Ventas Registradas — Lima Total (diario + forecast)',
                           C_AZUL, C_NARANJA, '2E86AB')
        sheet_mensual_excel(wb, 'Ventas Mensuales', h, fc,
                            'Ventas Mensuales — Lima Total',
                            C_AZUL)

    sheet_comparativo_excel(wb, resultados, 'altas')
    sheet_comparativo_excel(wb, resultados, 'ventas')

    nombre_archivo = f"Forecast_Aliv_{date.today().strftime('%Y-%m-%d')}.xlsx"
    ruta_salida = os.path.join(os.path.dirname(__file__), nombre_archivo)
    wb.save(ruta_salida)
    print(f"  Guardado: {ruta_salida}\n")

    # 4. Resumen en consola
    print("[ 4/4 ] Resumen de predicciones mensuales:")
    print(f"  {'Mes':<10} {'Altas Tot':>10} {'Alt.Vert':>10} {'Alt.Horiz':>10}  {'Ventas Tot':>11} {'Vta.Vert':>10} {'Vta.Horiz':>11}")
    print(f"  {'-'*9} {'-'*10} {'-'*10} {'-'*10}  {'-'*11} {'-'*10} {'-'*11}")

    def proyeccion_mes(key, mes):
        if key not in resultados:
            return '-'
        h, fc = resultados[key]
        df_m = forecast_mensual(h, fc)
        row = df_m[df_m['mes'] == mes]
        if row.empty:
            return '-'
        v = row.iloc[0]['proyeccion']
        return f"{int(round(v)):,}" if pd.notna(v) else '-'

    hoy = pd.Timestamp(date.today())
    for i in range(6):
        mes = pd.Period(hoy + pd.DateOffset(months=i), 'M')
        sufijo = ' (parcial)' if i == 0 else (' [FC]' if i > 0 else '')
        label = str(mes) + sufijo
        print(f"  {label:<19} {proyeccion_mes('altas_total', mes):>10} "
              f"{proyeccion_mes('altas_vert', mes):>10} "
              f"{proyeccion_mes('altas_horiz', mes):>10}  "
              f"{proyeccion_mes('ventas_total', mes):>11} "
              f"{proyeccion_mes('ventas_vert', mes):>10} "
              f"{proyeccion_mes('ventas_horiz', mes):>11}")

    print(f"\n[OK] Forecast completado: {nombre_archivo}")


if __name__ == '__main__':
    main()
